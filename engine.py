# -*- coding: utf-8 -*-
import openpyxl
import io
import os

def parse_html_wip_vcs(file_like_or_path):
    """
    Parses HTML-based tables or Excel-based WIP files dynamically
    to extract VEHICLE CODE and PLATE PUNCH DATE.
    """
    import pandas as pd
    import io
    
    try:
        # Read prefix/content to detect format
        if hasattr(file_like_or_path, 'read'):
            if hasattr(file_like_or_path, 'seek'):
                file_like_or_path.seek(0)
            content = file_like_or_path.read()
            if hasattr(file_like_or_path, 'seek'):
                file_like_or_path.seek(0)
                
            prefix = content[:300].lower()
            if b'<table' in prefix or b'<style' in prefix or b'<html' in prefix:
                # HTML parser
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(content, 'html.parser')
                rows = []
                for tr in soup.find_all('tr'):
                    cells = [td.get_text(strip=True) for td in tr.find_all(['td', 'th'])]
                    if cells:
                        rows.append(cells)
                if len(rows) < 2:
                    return []
                headers = rows[0]
                data = rows[1:]
                df = pd.DataFrame(data, columns=headers)
            else:
                # Excel reader
                filename = getattr(file_like_or_path, 'name', '').lower()
                if filename.endswith('.xlsb'):
                    df = pd.read_excel(io.BytesIO(content), engine='pyxlsb')
                else:
                    try:
                        df = pd.read_excel(io.BytesIO(content))
                    except Exception:
                        try:
                            dfs = pd.read_html(io.BytesIO(content), header=0)
                            df = dfs[0]
                        except Exception as e_inner:
                            raise ValueError(f"Could not read excel or HTML file: {e_inner}")
        else:
            # File path
            with open(file_like_or_path, 'rb') as f:
                prefix = f.read(300).lower()
                
            if b'<table' in prefix or b'<style' in prefix or b'<html' in prefix:
                from bs4 import BeautifulSoup
                with open(file_like_or_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                soup = BeautifulSoup(content, 'html.parser')
                rows = []
                for tr in soup.find_all('tr'):
                    cells = [td.get_text(strip=True) for td in tr.find_all(['td', 'th'])]
                    if cells:
                        rows.append(cells)
                if len(rows) < 2:
                    return []
                headers = rows[0]
                data = rows[1:]
                df = pd.DataFrame(data, columns=headers)
            else:
                filename = str(file_like_or_path).lower()
                if filename.endswith('.xlsb'):
                    df = pd.read_excel(file_like_or_path, engine='pyxlsb')
                else:
                    try:
                        df = pd.read_excel(file_like_or_path)
                    except Exception:
                        try:
                            dfs = pd.read_html(file_like_or_path, header=0)
                            df = dfs[0]
                        except Exception as e_inner:
                            raise ValueError(f"Could not read excel or HTML file: {e_inner}")
    except Exception as e:
        print(f"Failed to read WIP file: {e}")
        return []
        
    if df.empty:
        return []
        
    vc_col = None
    date_col = None
    for col in df.columns:
        col_clean = str(col).strip().upper()
        if 'VEHICLE CODE' in col_clean or 'VC' in col_clean:
            vc_col = col
        if 'PLATE PUNCH DATE' in col_clean or 'PUNCH' in col_clean:
            date_col = col
            
    if vc_col is None:
        vc_col = df.columns[4] if len(df.columns) > 4 else None
    if date_col is None:
        date_col = df.columns[8] if len(df.columns) > 8 else None
        
    # Detect description and color columns
    desc_col = None
    color_col = None
    for col in df.columns:
        col_clean = str(col).strip().upper()
        if 'SALES DESCRIPTION' in col_clean or 'DESC' in col_clean:
            desc_col = col
        if 'COLOR' in col_clean or 'COLOUR' in col_clean:
            color_col = col
            
    records = []
    for idx, row in df.iterrows():
        vc_val = row[vc_col] if vc_col is not None else None
        date_val = row[date_col] if date_col is not None else None
        desc_val = row[desc_col] if desc_col is not None else ''
        color_val = row[color_col] if color_col is not None else ''
        
        vc_str = str(vc_val).strip().upper() if vc_val is not None else ''
        desc_str = str(desc_val).strip().upper()
        color_str = str(color_val).strip().upper()
        
        # Apply Pune plant exclusions
        if vc_str.startswith('5442') or 'ALTROZ' in desc_str or color_str == '000' or '000' in color_str:
            continue
            
        if vc_val is not None and str(vc_val).strip() != '':
            records.append({
                'vc': vc_str,
                'date_str': str(date_val).strip() if date_val is not None else ''
            })
            
    return records

def initialize_tcf1_workbook(file_path_or_stream):
    """
    Initializes the TCF-1 workbook for the new day's planning run:
    1. Copies yesterday's Day 1 date and values (col H) to Today Plan date and values (col D) dynamically.
    2. Clears 'biw Next 2 days' columns B, C, and D.
    3. Clears 'VIN & Expected VIN' (or 'VIN & Expected VIN ') columns A–D and F–I from row 3 downwards.
    4. Clears 'Day 1 Sequence' columns B, C, D from row 3 downwards.
    """
    # 1. Handle dual load (data-only for values, formula-only for writing)
    if hasattr(file_path_or_stream, 'read'):
        if hasattr(file_path_or_stream, 'seek'):
            file_path_or_stream.seek(0)
        content = file_path_or_stream.read()
        if hasattr(file_path_or_stream, 'seek'):
            file_path_or_stream.seek(0)
        wb_val = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    else:
        wb_val = openpyxl.load_workbook(file_path_or_stream, data_only=True)
        wb = openpyxl.load_workbook(file_path_or_stream, data_only=False)
        
    # Copy H (Day 1) values to D (Today Plan) in Plan Summary
    if 'Plan Summary' in wb.sheetnames and 'Plan Summary' in wb_val.sheetnames:
        ws_val = wb_val['Plan Summary']
        ws = wb['Plan Summary']
        
        # Copy yesterday's Day 1 date (col H) to Today's Today Plan date (col D)
        yesterday_day1_date = ws_val.cell(row=2, column=8).value
        ws.cell(row=2, column=4).value = yesterday_day1_date
        
        # Calculate next 4 planning days skipping Sundays
        import datetime
        start_date = yesterday_day1_date
        if isinstance(start_date, datetime.datetime):
            start_date = start_date.date()
            
        def get_next_planning_days_t1(s_date, count=4):
            days = []
            curr = s_date
            while len(days) < count:
                curr += datetime.timedelta(days=1)
                if curr.weekday() == 6: # Sunday
                    continue
                days.append(curr)
            return days
            
        if start_date is not None and not isinstance(start_date, str):
            planning_days = get_next_planning_days_t1(start_date, 4)
            ws.cell(row=2, column=8).value = planning_days[0]
            ws.cell(row=2, column=9).value = planning_days[1]
            ws.cell(row=2, column=10).value = planning_days[2]
            ws.cell(row=2, column=11).value = planning_days[3]
        
        # Row-agnostic copy range
        for row in range(4, ws.max_row + 1):
            col1_val = str(ws.cell(row=row, column=1).value or '').strip().upper()
            if 'TOTAL' in col1_val or col1_val == '':
                continue
            # Read evaluated value from col H (8)
            day1_val = ws_val.cell(row=row, column=8).value
            # Write evaluated value to col D (4)
            ws.cell(row=row, column=4).value = day1_val
            
    # 2. biw Next 2 days: Clear B, C & D
    if 'biw Next 2 days' in wb.sheetnames:
        ws_biw = wb['biw Next 2 days']
        for r in range(2, ws_biw.max_row + 1):
            ws_biw.cell(row=r, column=2).value = None
            ws_biw.cell(row=r, column=3).value = None
            ws_biw.cell(row=r, column=4).value = None
            
    # 3. Stage wise Float: Clear B, D, E, F, G, H (backward compatibility)
    if 'Stage wise Float' in wb.sheetnames:
        ws_float = wb['Stage wise Float']
        cols_to_clear = [2, 4, 5, 6, 7, 8]  # B, D, E, F, G, H
        for r in range(2, ws_float.max_row + 1):
            for col in cols_to_clear:
                ws_float.cell(row=r, column=col).value = None
                
    # 4. Day 1 Sequence: Clear B, C, D
    if 'Day 1 Sequence' in wb.sheetnames:
        ws_seq = wb['Day 1 Sequence']
        cols_seq = [2, 3, 4]  # B, C, D
        for r in range(3, ws_seq.max_row + 1):
            for col in cols_seq:
                ws_seq.cell(row=r, column=col).value = None
                
    # 5. Vin Plan Check: Clear F (Expected to Drop) (backward compatibility)
    if 'Vin Plan Check' in wb.sheetnames:
        ws_check = wb['Vin Plan Check']
        for r in range(4, ws_check.max_row + 1):
            ws_check.cell(row=r, column=6).value = None
            
    # 6. VIN & Expected VIN: Clear columns A-D and F-I from Row 3 downwards
    expected_sheet_name = 'VIN & Expected VIN'
    if 'VIN & Expected VIN ' in wb.sheetnames:
        expected_sheet_name = 'VIN & Expected VIN '
    if expected_sheet_name in wb.sheetnames:
        ws_expected = wb[expected_sheet_name]
        for r in range(3, ws_expected.max_row + 1):
            for col in [1, 2, 3, 4, 6, 7, 8, 9]:
                ws_expected.cell(row=r, column=col).value = None
            
    return wb

def initialize_tcf2_workbook(file_path_or_stream):
    """
    Initializes the TCF-2 workbook for the new day's planning run:
    1. Copies yesterday's Day 1 date and values (col I) to Days Plan date and values (col E) dynamically.
    2. Clears 'biw Next 2 days' columns B, C, and D.
    3. Clears 'VIN & Expected VIN' (or 'VIN & Expected VIN ') columns A–D and F–I from row 3 downwards.
    4. Clears 'Day 1 Sequence' columns B, C, D from row 3 downwards.
    """
    # 1. Handle dual load (data-only for values, formula-only for writing)
    if hasattr(file_path_or_stream, 'read'):
        if hasattr(file_path_or_stream, 'seek'):
            file_path_or_stream.seek(0)
        content = file_path_or_stream.read()
        if hasattr(file_path_or_stream, 'seek'):
            file_path_or_stream.seek(0)
        wb_val = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    else:
        wb_val = openpyxl.load_workbook(file_path_or_stream, data_only=True)
        wb = openpyxl.load_workbook(file_path_or_stream, data_only=False)
        
    # Copy I (Day 1) values to E (Days Plan) in Plan Summary
    if 'Plan Summary' in wb.sheetnames and 'Plan Summary' in wb_val.sheetnames:
        ws_val = wb_val['Plan Summary']
        ws = wb['Plan Summary']
        
        # Copy yesterday's Day 1 date (col I) to Today's Today Plan date (col E)
        yesterday_day1_date = ws_val.cell(row=2, column=9).value
        ws.cell(row=2, column=5).value = yesterday_day1_date
        
        # Calculate next 4 planning days skipping Sundays
        import datetime
        start_date = yesterday_day1_date
        if isinstance(start_date, datetime.datetime):
            start_date = start_date.date()
            
        def get_next_planning_days_t2(s_date, count=4):
            days = []
            curr = s_date
            while len(days) < count:
                curr += datetime.timedelta(days=1)
                if curr.weekday() == 6: # Sunday
                    continue
                days.append(curr)
            return days
            
        if start_date is not None and not isinstance(start_date, str):
            planning_days = get_next_planning_days_t2(start_date, 4)
            ws.cell(row=2, column=9).value = planning_days[0]
            ws.cell(row=2, column=10).value = planning_days[1]
            ws.cell(row=2, column=11).value = planning_days[2]
            ws.cell(row=2, column=12).value = planning_days[3] # Column L
        
        # Row-agnostic copy range
        for row in range(4, ws.max_row + 1):
            col2_val = str(ws.cell(row=row, column=2).value or '').strip().upper()
            col3_val = str(ws.cell(row=row, column=3).value or '').strip().upper()
            if 'TOTAL' in col2_val or 'TOTAL' in col3_val or (col2_val == '' and col3_val == ''):
                continue
            # Read evaluated value from col I (9)
            day1_val = ws_val.cell(row=row, column=9).value
            # Write evaluated value to col E (5)
            ws.cell(row=row, column=5).value = day1_val
            
    # 2. biw Next 2 days: Clear B, C & D
    if 'biw Next 2 days' in wb.sheetnames:
        ws_biw = wb['biw Next 2 days']
        for r in range(2, ws_biw.max_row + 1):
            ws_biw.cell(row=r, column=2).value = None
            ws_biw.cell(row=r, column=3).value = None
            ws_biw.cell(row=r, column=4).value = None
            
    # 3. Stage wise float: Clear B, D, E, F, G, H (backward compatibility)
    sheet_name = 'Stage wise float'
    if 'Stage wise float ' in wb.sheetnames:
        sheet_name = 'Stage wise float '
    if sheet_name in wb.sheetnames:
        ws_float = wb[sheet_name]
        cols_to_clear = [2, 4, 5, 6, 7, 8]  # B, D, E, F, G, H
        for r in range(2, ws_float.max_row + 1):
            for col in cols_to_clear:
                ws_float.cell(row=r, column=col).value = None
                
    # 4. Day 1 Sequence: Clear B, C, D
    if 'Day 1 Sequence' in wb.sheetnames:
        ws_seq = wb['Day 1 Sequence']
        cols_seq = [2, 3, 4]  # B, C, D
        for r in range(3, ws_seq.max_row + 1):
            for col in cols_seq:
                ws_seq.cell(row=r, column=col).value = None
                
    # 5. Vin plan check: Clear G (Expected Drop) (backward compatibility)
    sheet_name_check = 'Vin plan check'
    if 'Vin Plan Check' in wb.sheetnames:
        sheet_name_check = 'Vin Plan Check'
    if sheet_name_check in wb.sheetnames:
        ws_check = wb[sheet_name_check]
        for r in range(3, ws_check.max_row + 1):
            ws_check.cell(row=r, column=7).value = None
            
    # 6. VIN & Expected VIN: Clear columns A-D and F-I from Row 3 downwards
    expected_sheet_name = 'VIN & Expected VIN'
    if 'VIN & Expected VIN ' in wb.sheetnames:
        expected_sheet_name = 'VIN & Expected VIN '
    if expected_sheet_name in wb.sheetnames:
        ws_expected = wb[expected_sheet_name]
        for r in range(3, ws_expected.max_row + 1):
            for col in [1, 2, 3, 4, 6, 7, 8, 9]:
                ws_expected.cell(row=r, column=col).value = None
            
    return wb

def update_today_vin_generation(wb, vin_list_file_or_stream, track):
    """
    Parses the Today's VIN List file and writes the VINs and details (Sr No, Vehicle Code,
    Sales Description, Color) to Table 1 (Columns A to D) of the 'VIN & Expected VIN' sheet,
    maintaining backwards-compatibility with the old float sheet.
    """
    import pandas as pd
    import io
    
    # 1. Read the VIN list file using pandas
    try:
        # Check if it's a file stream or a file path
        if hasattr(vin_list_file_or_stream, 'read'):
            # It's a stream (e.g. UploadedFile)
            if hasattr(vin_list_file_or_stream, 'seek'):
                vin_list_file_or_stream.seek(0)
            content = vin_list_file_or_stream.read()
            # Reset seek position just in case
            if hasattr(vin_list_file_or_stream, 'seek'):
                vin_list_file_or_stream.seek(0)
            
            # Check for HTML signature (for .xls files exported as HTML tables)
            prefix = content[:300].lower()
            if b'<table' in prefix or b'<style' in prefix or b'<html' in prefix:
                dfs = pd.read_html(io.BytesIO(content), header=0)
                df = dfs[0]
            else:
                # Determine format based on name attribute
                filename = getattr(vin_list_file_or_stream, 'name', '').lower()
                if filename.endswith('.xlsb'):
                    df = pd.read_excel(io.BytesIO(content), engine='pyxlsb')
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        df = pd.read_excel(io.BytesIO(content))
                    except Exception:
                        try:
                            dfs = pd.read_html(io.BytesIO(content), header=0)
                            df = dfs[0]
                        except Exception as e_inner:
                            raise ValueError(f"Could not read excel or HTML file: {e_inner}")
                else:
                    # Fallback to CSV
                    try:
                        df = pd.read_csv(io.BytesIO(content), encoding='utf-8')
                    except Exception:
                        df = pd.read_csv(io.BytesIO(content), encoding='latin1')
        else:
            # It's a file path
            with open(vin_list_file_or_stream, 'rb') as f:
                prefix = f.read(300).lower()
            
            if b'<table' in prefix or b'<style' in prefix or b'<html' in prefix:
                dfs = pd.read_html(vin_list_file_or_stream, header=0)
                df = dfs[0]
            else:
                filename = str(vin_list_file_or_stream).lower()
                if filename.endswith('.xlsb'):
                    df = pd.read_excel(vin_list_file_or_stream, engine='pyxlsb')
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        df = pd.read_excel(vin_list_file_or_stream)
                    except Exception:
                        try:
                            dfs = pd.read_html(vin_list_file_or_stream, header=0)
                            df = dfs[0]
                        except Exception as e_inner:
                            raise ValueError(f"Could not read excel or HTML file: {e_inner}")
                else:
                    df = pd.read_csv(vin_list_file_or_stream)
    except Exception as e:
        raise ValueError(f"Failed to read VIN list file: {e}")
        
    # 2. Extract records from the dataframe
    records = []
    if not df.empty:
        # Find columns dynamically
        def find_col_in_df(patterns):
            for pattern in patterns:
                for col in df.columns:
                    if pattern.lower() in str(col).lower():
                        return col
            return None
            
        vc_col = find_col_in_df(['vehicle code', 'vehicle_code', 'vc', 'vin', 'chassis'])
        desc_col = find_col_in_df(['sales description', 'sales_description', 'desc', 'model'])
        color_col = find_col_in_df(['color', 'colour', 'clr'])
        
        # Fallbacks for VC column if not found
        if vc_col is None:
            for col in df.columns:
                if 'vehicle' in str(col).lower():
                    vc_col = col
                    break
        if vc_col is None:
            for col in df.columns:
                col_lower = str(col).lower()
                if 'vin' in col_lower or 'chassis' in col_lower:
                    vc_col = col
                    break
        if vc_col is None:
            keywords = ['serial', 'number']
            for col in df.columns:
                col_str = str(col).lower()
                if any(k in col_str for k in keywords):
                    vc_col = col
                    break
        if vc_col is None:
            for col in df.columns:
                non_nulls = df[col].dropna()
                if not non_nulls.empty:
                    string_lengths = non_nulls.astype(str).str.len()
                    valid_len_pct = ((string_lengths >= 8) & (string_lengths <= 20)).mean()
                    if valid_len_pct >= 0.5:
                        vc_col = col
                        break
        if vc_col is None:
            vc_col = df.columns[0]
            
        # Parse rows
        for idx, row in df.iterrows():
            vc_val = str(row[vc_col]).strip() if vc_col is not None and pd.notna(row[vc_col]) else ""
            desc_val = str(row[desc_col]).strip() if desc_col is not None and pd.notna(row[desc_col]) else ""
            color_val = str(row[color_col]).strip() if color_col is not None and pd.notna(row[color_col]) else ""
            
            if vc_val != "":
                records.append({
                    'vc': vc_val.upper(),
                    'desc': desc_val,
                    'color': color_val
                })
                
    # 3. Write records to the sheet VIN & Expected VIN
    expected_sheet_name = 'VIN & Expected VIN'
    if 'VIN & Expected VIN ' in wb.sheetnames:
        expected_sheet_name = 'VIN & Expected VIN '
        
    if expected_sheet_name in wb.sheetnames:
        ws = wb[expected_sheet_name]
        for idx, rec in enumerate(records):
            row_idx = idx + 3
            ws.cell(row=row_idx, column=1).value = idx + 1  # Sr No
            ws.cell(row=row_idx, column=2).value = rec['vc']
            ws.cell(row=row_idx, column=3).value = rec['desc']
            ws.cell(row=row_idx, column=4).value = rec['color']
            
    # 4. Backwards-compatibility: write only VCs to old float sheet if present
    old_sheet_name = 'Stage wise Float' if track == 'TCF1' else 'Stage wise float'
    if old_sheet_name not in wb.sheetnames and track == 'TCF2':
        if 'Stage wise float ' in wb.sheetnames:
            old_sheet_name = 'Stage wise float '
    if old_sheet_name in wb.sheetnames:
        ws_old = wb[old_sheet_name]
        for idx, rec in enumerate(records):
            row_idx = idx + 2
            ws_old.cell(row=row_idx, column=2).value = rec['vc']
            
    return wb

def update_paint_float_data(wb, paint_float_file_or_stream, track, expected_qty=None,
                            wip_files=None, yest_plan_file_or_stream=None,
                            next_3days_biw_plan=None, daily_capacities=None,
                            sub_limits_1=None, sub_limits_2=None):
    """
    Parses the Paint Float Report, updates Table 2 of 'VIN & Expected VIN' (or 'VIN & Expected VIN ') sheet,
    populates the old 'Stage wise Float' and 'Vin Plan Check' sheets (if present) for backwards-compatibility,
    and runs the 4-day plan sequencing for both tracks.
    """
    import pandas as pd
    import io
    
    # 1. Read the Paint Float Report using pandas
    try:
        if hasattr(paint_float_file_or_stream, 'read'):
            if hasattr(paint_float_file_or_stream, 'seek'):
                paint_float_file_or_stream.seek(0)
            content = paint_float_file_or_stream.read()
            if hasattr(paint_float_file_or_stream, 'seek'):
                paint_float_file_or_stream.seek(0)
            
            prefix = content[:300].lower()
            if b'<table' in prefix or b'<style' in prefix or b'<html' in prefix:
                dfs = pd.read_html(io.BytesIO(content), header=0)
                df = dfs[0]
            else:
                filename = getattr(paint_float_file_or_stream, 'name', '').lower()
                if filename.endswith('.xlsb'):
                    df = pd.read_excel(io.BytesIO(content), engine='pyxlsb')
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        df = pd.read_excel(io.BytesIO(content))
                    except Exception:
                        try:
                            dfs = pd.read_html(io.BytesIO(content), header=0)
                            df = dfs[0]
                        except Exception as e_inner:
                            raise ValueError(f"Could not read excel or HTML file: {e_inner}")
                else:
                    try:
                        df = pd.read_csv(io.BytesIO(content), encoding='utf-8')
                    except Exception:
                        df = pd.read_csv(io.BytesIO(content), encoding='latin1')
        else:
            with open(paint_float_file_or_stream, 'rb') as f:
                prefix = f.read(300).lower()
            
            if b'<table' in prefix or b'<style' in prefix or b'<html' in prefix:
                dfs = pd.read_html(paint_float_file_or_stream, header=0)
                df = dfs[0]
            else:
                filename = str(paint_float_file_or_stream).lower()
                if filename.endswith('.xlsb'):
                    df = pd.read_excel(paint_float_file_or_stream, engine='pyxlsb')
                elif filename.endswith(('.xlsx', '.xls')):
                    try:
                        df = pd.read_excel(paint_float_file_or_stream)
                    except Exception:
                        try:
                            dfs = pd.read_html(paint_float_file_or_stream, header=0)
                            df = dfs[0]
                        except Exception as e_inner:
                            raise ValueError(f"Could not read excel or HTML file: {e_inner}")
                else:
                    df = pd.read_csv(paint_float_file_or_stream)
    except Exception as e:
        raise ValueError(f"Failed to read Paint Float Report file: {e}")
        
    if df.empty:
        return wb
        
    # Helper to find column dynamically
    def find_col(patterns):
        for pattern in patterns:
            for col in df.columns:
                if pattern.lower() in str(col).lower():
                    return col
        return None
        
    # Find key columns
    vin_col = find_col(['vehicle code', 'vehicle_code', 'vc', 'vin', 'chassis'])
    shop_col = find_col(['shop'])
    
    if not vin_col or not shop_col:
        raise ValueError("Paint Float Report must contain 'VEHICLE CODE' and 'SHOP' columns.")
        
    # Filter by track/shop (TCF1 or TCF2)
    shop_val = 'TCF1' if track == 'TCF1' else 'TCF2'
    df_filtered = df[df[shop_col].astype(str).str.strip().str.upper() == shop_val.upper()]
    
    # Stage mappings
    stage_mappings = [
        (['pbs lift', 'pbs', 'pbs_lift'], 4),        # Col D (4)
        (['topcoat', 'top coat', 'top_coat'], 5),    # Col E (5)
        (['sealant', 'seal'], 6),                    # Col F (6)
        (['ptced', 'ptc'], 7),                       # Col G (7)
        (['biw lifting', 'biw lift', 'biw_lifting'], 8) # Col H (8)
    ]
    
    available_df = df_filtered.copy()
    stage_vcs_dict = {}
    
    for patterns, col_idx in stage_mappings:
        src_col = find_col(patterns)
        if src_col and not available_df.empty:
            valid_rows = available_df[
                available_df[src_col].notna() & 
                (available_df[src_col].astype(str).str.strip() != '')
            ]
            stage_vins = valid_rows[vin_col].dropna().astype(str).str.strip().tolist()
            stage_vcs_dict[col_idx] = stage_vins
            available_df = available_df.drop(valid_rows.index)
        else:
            stage_vcs_dict[col_idx] = []
            
    # Load VIN Color Plan to check Pune plant exclusions
    color_sheet_name = 'VIN Color Plan' if 'VIN Color Plan' in wb.sheetnames else 'Vin Color Plan'
    if color_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{color_sheet_name}' not found in the workbook.")
    color_ws = wb[color_sheet_name]
    
    data_start_row = 3 if track == 'TCF1' else 2
    
    vc_lookup = {}
    for r_color in range(data_start_row, color_ws.max_row + 1):
        vc_val = color_ws.cell(row=r_color, column=3).value  # Col C is Vehicle Code
        if vc_val is not None:
            vc_str = str(vc_val).strip().upper()
            sales_desc = color_ws.cell(row=r_color, column=5).value  # Col E is Description
            colour_val = color_ws.cell(row=r_color, column=6).value  # Col F is Clr
            short_vc = color_ws.cell(row=r_color, column=4).value  # Col D is Short VC
            vc_lookup[vc_str] = (sales_desc, colour_val, short_vc)

    # Flatten stages column-by-column and apply exclusions
    all_stage_vcs = []
    for col_idx in [4, 5, 6, 7, 8]:
        for vin in stage_vcs_dict[col_idx]:
            val_clean = vin.upper()
            
            # Retrieve from lookup to check Pune plant exclusions
            desc_val, col_val, svc_val = vc_lookup.get(val_clean, (None, None, ""))
            desc_str = str(desc_val).strip().upper() if desc_val is not None else ''
            color_str = str(col_val).strip().upper() if col_val is not None else ''
            
            if val_clean.startswith('5442') or 'ALTROZ' in desc_str or color_str == '000' or '000' in color_str:
                continue
                
            all_stage_vcs.append(val_clean)

    # Resolve target quantity
    if expected_qty is not None:
        try:
            expected_qty = int(expected_qty)
        except Exception:
            expected_qty = 500 if track == 'TCF1' else 100
    else:
        expected_qty = 500 if track == 'TCF1' else 100

    # Pick expected drops
    picked_vcs = all_stage_vcs[:expected_qty]
    remaining_vcs = all_stage_vcs[expected_qty:]

    # Write expected drops details to VIN & Expected VIN sheet
    expected_sheet_name = 'VIN & Expected VIN'
    if 'VIN & Expected VIN ' in wb.sheetnames:
        expected_sheet_name = 'VIN & Expected VIN '
    if expected_sheet_name in wb.sheetnames:
        ws_expected = wb[expected_sheet_name]
        for idx, vc in enumerate(picked_vcs):
            row_idx = idx + 3
            ws_expected.cell(row=row_idx, column=6).value = idx + 1 # Sr No
            ws_expected.cell(row=row_idx, column=7).value = vc
            
            desc_val, col_val, svc_val = vc_lookup.get(vc, ("", "", ""))
            ws_expected.cell(row=row_idx, column=8).value = desc_val
            ws_expected.cell(row=row_idx, column=9).value = col_val

    # Backwards-compatibility: write stages to old Stage wise Float sheet if present
    old_float_name = 'Stage wise Float' if track == 'TCF1' else 'Stage wise float'
    if old_float_name not in wb.sheetnames and track == 'TCF2':
        if 'Stage wise float ' in wb.sheetnames:
            old_float_name = 'Stage wise float '
    if old_float_name in wb.sheetnames:
        ws_float_old = wb[old_float_name]
        from openpyxl.styles import PatternFill
        light_green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # Write back to stages columns D-H
        picked_set = set(picked_vcs)
        for col_idx in [4, 5, 6, 7, 8]:
            stage_vins = stage_vcs_dict[col_idx]
            for idx, vin in enumerate(stage_vins):
                row_idx = idx + 2
                ws_float_old.cell(row=row_idx, column=col_idx).value = vin
                
                # Exclude from highlight if Pune plant excluded
                val_clean = vin.upper()
                desc_val, col_val, svc_val = vc_lookup.get(val_clean, (None, None, ""))
                desc_str = str(desc_val).strip().upper() if desc_val is not None else ''
                color_str = str(col_val).strip().upper() if col_val is not None else ''
                if val_clean.startswith('5442') or 'ALTROZ' in desc_str or color_str == '000' or '000' in color_str:
                    ws_float_old.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type=None)
                    continue
                
                if val_clean in picked_set:
                    ws_float_old.cell(row=row_idx, column=col_idx).fill = light_green_fill
                else:
                    ws_float_old.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type=None)

    # Backwards-compatibility: Write counts to old Vin Plan Check / Vin plan check if present
    from collections import Counter
    vc_counts = Counter(picked_vcs)
    check_sheet_name = 'Vin Plan Check' if track == 'TCF1' else 'Vin plan check'
    if check_sheet_name not in wb.sheetnames and track == 'TCF2':
        if 'Vin Plan Check' in wb.sheetnames:
            check_sheet_name = 'Vin Plan Check'
    if check_sheet_name in wb.sheetnames:
        check_ws = wb[check_sheet_name]
        start_row = 4 if track == 'TCF1' else 3
        vc_col_idx = 2 if track == 'TCF1' else 3
        drop_col_idx = 6 if track == 'TCF1' else 7
        
        r = start_row
        while True:
            sr_no_val = check_ws.cell(row=r, column=1).value
            if r_val := check_ws.cell(row=r, column=1).value:
                sr_no_val = r_val
            else:
                break
            try:
                int(float(str(sr_no_val).strip()))
            except ValueError:
                break
                
            vc_val = check_ws.cell(row=r, column=vc_col_idx).value
            if vc_val is not None:
                vc_str = str(vc_val).strip().upper()
                count = vc_counts.get(vc_str, 0)
                check_ws.cell(row=r, column=drop_col_idx).value = count
            else:
                check_ws.cell(row=r, column=drop_col_idx).value = 0
            r += 1

    # Day 1 Sequence sheet setup
    day1_sheet_name = 'Day 1 Sequence'
    if day1_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{day1_sheet_name}' not found in the workbook.")
    day1_ws = wb[day1_sheet_name]
    
    # 3-Day Plan sequencing
    if wip_files is not None:
        wip_records = []
        if isinstance(wip_files, list):
            for wip_file in wip_files:
                if wip_file is not None:
                    wip_records.extend(parse_html_wip_vcs(wip_file))
        elif wip_files is not None:
            wip_records.extend(parse_html_wip_vcs(wip_files))
            
        for rec in wip_records:
            try:
                rec['datetime'] = pd.to_datetime(rec['date_str'], format='%d/%m/%Y %I:%M:%S %p')
            except Exception:
                try:
                    rec['datetime'] = pd.to_datetime(rec['date_str'])
                except Exception:
                    rec['datetime'] = pd.Timestamp.min
        wip_records.sort(key=lambda x: x['datetime'])
        wip_vcs = [(x['vc'], 'light_blue') for x in wip_records]
        
        master_yest_queue = []
        if yest_plan_file_or_stream is not None:
            if hasattr(yest_plan_file_or_stream, 'read'):
                if hasattr(yest_plan_file_or_stream, 'seek'):
                    yest_plan_file_or_stream.seek(0)
                content = yest_plan_file_or_stream.read()
                if hasattr(yest_plan_file_or_stream, 'seek'):
                    yest_plan_file_or_stream.seek(0)
                yest_wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            else:
                yest_wb = openpyxl.load_workbook(yest_plan_file_or_stream, data_only=True)
                
            if 'biw Next 2 days' in yest_wb.sheetnames:
                yest_ws = yest_wb['biw Next 2 days']
                cols_to_read = [2, 3, 4]  # Read all columns B, C, D
                for col_idx in cols_to_read:
                    for r in range(2, yest_ws.max_row + 1):
                        val = yest_ws.cell(row=r, column=col_idx).value
                        if val is not None and str(val).strip() != '':
                            master_yest_queue.append((str(val).strip().upper(), 'light_yellow'))
                            
        if next_3days_biw_plan is not None:
            try:
                if hasattr(next_3days_biw_plan, 'read'):
                    if hasattr(next_3days_biw_plan, 'seek'):
                        next_3days_biw_plan.seek(0)
                    biw_content = next_3days_biw_plan.read()
                    if hasattr(next_3days_biw_plan, 'seek'):
                        next_3days_biw_plan.seek(0)
                    biw_wb = openpyxl.load_workbook(io.BytesIO(biw_content), data_only=True)
                else:
                    biw_wb = openpyxl.load_workbook(next_3days_biw_plan, data_only=True)
                
                if track == 'TCF1':
                    sheets_to_read = ['Punch Seq', 'NOVA']
                else:
                    sheets_to_read = ['Q5', 'TGDI ', 'Eturna']
                
                for sh_pat in sheets_to_read:
                    target_sh = None
                    for name in biw_wb.sheetnames:
                        if name.strip().upper() == sh_pat.strip().upper():
                            target_sh = name
                            break
                    if target_sh is not None:
                        ws_biw = biw_wb[target_sh]
                        headers = [str(c.value).strip().upper() if c.value is not None else '' for c in ws_biw[1]]
                        vc_col_idx = None
                        for idx, h in enumerate(headers):
                            if any(pat in h for pat in ['ITEMCODE', 'ITEM CODE', 'ITEM_CODE', 'VEHICLE CODE', 'VEHICLE_CODE', 'VC']):
                                vc_col_idx = idx + 1
                                break
                        if vc_col_idx is None:
                            vc_col_idx = 2 if ws_biw.max_column >= 2 else 1
                        
                        for r in range(2, ws_biw.max_row + 1):
                            val = ws_biw.cell(row=r, column=vc_col_idx).value
                            if val is not None:
                                vc_str = str(val).strip().upper()
                                if vc_str != '':
                                    desc_val, col_val, svc_val = vc_lookup.get(vc_str, (None, None, ""))
                                    desc_str = str(desc_val).strip().upper() if desc_val is not None else ''
                                    color_str = str(col_val).strip().upper() if col_val is not None else ''
                                    
                                    if vc_str.startswith('5442') or 'ALTROZ' in desc_str or color_str == '000' or '000' in color_str:
                                        continue
                                    master_yest_queue.append((vc_str, 'light_pink'))
            except Exception as e_biw:
                print(f"Error reading Next 3-Days BIW Plan: {e_biw}")
                            
        def identify_vehicle_type(vc, desc, svc=""):
            vc_upper = str(vc).upper()
            desc_upper = str(desc or '').upper()
            svc_upper = str(svc or '').strip().upper()
            
            # TCF-1 Nova (EV)
            if svc_upper.startswith('5468') or 'PUNCH.EV' in desc_upper or 'PUNCH EV' in desc_upper or 'PUNCH.EV' in vc_upper or 'PUNCH EV' in vc_upper:
                return 'NOVA'
            # TCF-1 CNG
            if 'CNG' in desc_upper or 'CNG' in vc_upper or svc_upper.startswith('549718') or svc_upper.startswith('549722') or svc_upper.startswith('549728'):
                return 'CNG'
                
            # TCF-2 Eturna (EV)
            if svc_upper.startswith('5473') or svc_upper.startswith('5483') or 'HARRIER.EV' in desc_upper or 'HARRIER EV' in desc_upper or 'SAFARI.EV' in desc_upper or 'SAFARI EV' in desc_upper or 'HARRIER.EV' in vc_upper or 'SAFARI.EV' in vc_upper:
                return 'ETURNA'
            # TCF-2 TGDI (Petrol)
            if svc_upper.startswith('5478') or svc_upper.startswith('5479') or 'TGDI' in desc_upper or 'TGDI' in vc_upper:
                return 'TGDI'
            if ' BS6 P' in desc_upper or ' BS6 P2 P' in desc_upper or 'BS6 P2 P' in desc_upper:
                if 'BS6 D' not in desc_upper and 'P2 D' not in desc_upper:
                    return 'TGDI'
                    
            return 'OTHER'

        def build_day_sequence(indexed_queue, max_capacity, lim1_val, lim2_val, is_tcf1=True):
            if max_capacity <= 0:
                return [], indexed_queue, 0, 0
                
            ev_list = []
            c2_list = []
            other_list = []
            
            for orig_idx, item in indexed_queue:
                vc, color_key = item
                desc_val, col_val, svc_val = vc_lookup.get(vc, (None, None, ""))
                v_type = identify_vehicle_type(vc, desc_val, svc_val)
                
                if is_tcf1:
                    if v_type == 'NOVA':
                        ev_list.append((orig_idx, item))
                    elif v_type == 'CNG':
                        c2_list.append((orig_idx, item))
                    else:
                        other_list.append((orig_idx, item))
                else:
                    if v_type == 'ETURNA':
                        ev_list.append((orig_idx, item))
                    elif v_type == 'TGDI':
                        c2_list.append((orig_idx, item))
                    else:
                        other_list.append((orig_idx, item))
            
            rem_val = max(0, max_capacity - lim1_val - lim2_val)
            
            selected_ev = ev_list[:lim1_val]
            selected_c2 = c2_list[:lim2_val]
            selected_other = other_list[:rem_val]
            
            selected_combined = selected_ev + selected_c2 + selected_other
            selected_combined.sort(key=lambda x: x[0])
            
            day_list = [item for orig_idx, item in selected_combined]
            selected_orig_indices = set(orig_idx for orig_idx, item in selected_combined)
            postponed = [(orig_idx, item) for orig_idx, item in indexed_queue if orig_idx not in selected_orig_indices]
            
            c1_count = len(selected_ev)
            c2_count = len(selected_c2)
            
            return day_list, postponed, c1_count, c2_count

        picked_n, picked_c, picked_e, picked_t = 0, 0, 0, 0
        for vc in picked_vcs:
            desc_val, col_val, svc_val = vc_lookup.get(vc, (None, None, ""))
            v_type = identify_vehicle_type(vc, desc_val, svc_val)
            if v_type == 'NOVA':
                picked_n += 1
            elif v_type == 'CNG':
                picked_c += 1
            elif v_type == 'ETURNA':
                picked_e += 1
            elif v_type == 'TGDI':
                picked_t += 1

        master_queue = []
        master_queue.extend([(vc, 'light_green') for vc in remaining_vcs])
        master_queue.extend(wip_vcs)
        master_queue.extend(master_yest_queue)
        
        indexed_master_queue = list(enumerate(master_queue))
        
        is_tcf1 = (track == 'TCF1')
        if daily_capacities is None:
            caps = [900, 900, 900, 900] if is_tcf1 else [250, 250, 250, 250]
        else:
            caps = list(daily_capacities)
            expected_len = 4
            while len(caps) < expected_len:
                caps.append(900 if is_tcf1 else 250)
                
        if sub_limits_1 is None:
            lim1 = [160, 160, 160, 160]
        else:
            lim1 = list(sub_limits_1)
            expected_len = 4
            while len(lim1) < expected_len:
                lim1.append(160)
                
        if sub_limits_2 is None:
            lim2 = [350, 200, 350, 350] if is_tcf1 else [40, 40, 40, 40]
        else:
            lim2 = list(sub_limits_2)
            expected_len = 4
            while len(lim2) < expected_len:
                lim2.append(350 if is_tcf1 else 40)
        
        day1_list, remaining_queue, d1_c1, d1_c2 = build_day_sequence(indexed_master_queue, caps[0], lim1[0], lim2[0], is_tcf1)
        day2_list, remaining_queue, d2_c1, d2_c2 = build_day_sequence(remaining_queue, caps[1], lim1[1], lim2[1], is_tcf1)
        day3_list, remaining_queue, d3_c1, d3_c2 = build_day_sequence(remaining_queue, caps[2], lim1[2], lim2[2], is_tcf1)
        day4_list, remaining_queue, d4_c1, d4_c2 = build_day_sequence(remaining_queue, caps[3], lim1[3], lim2[3], is_tcf1)
            
        all_planned_items = day1_list + day2_list + day3_list + day4_list
        color_counts = {
            'light_green': 0,
            'light_blue': 0,
            'light_yellow': 0,
            'light_pink': 0
        }
        vc_lists = {
            'light_green': [],
            'light_blue': [],
            'light_yellow': [],
            'light_pink': []
        }
        for item in all_planned_items:
            vc, color_key = item
            if color_key in color_counts:
                color_counts[color_key] += 1
            if color_key in vc_lists:
                vc_lists[color_key].append(vc)
                
        if is_tcf1:
            wb.summary_counts = {
                'picked': {'nova': picked_n, 'cng': picked_c, 'eturna': 0, 'tgdi': 0},
                'day1': {'nova': d1_c1, 'cng': d1_c2, 'eturna': 0, 'tgdi': 0},
                'day2': {'nova': d2_c1, 'cng': d2_c2, 'eturna': 0, 'tgdi': 0},
                'day3': {'nova': d3_c1, 'cng': d3_c2, 'eturna': 0, 'tgdi': 0},
                'day4': {'nova': d4_c1, 'cng': d4_c2, 'eturna': 0, 'tgdi': 0},
                'targets': {
                    'day1': {'nova': lim1[0], 'cng': lim2[0], 'eturna': 0, 'tgdi': 0},
                    'day2': {'nova': lim1[1], 'cng': lim2[1], 'eturna': 0, 'tgdi': 0},
                    'day3': {'nova': lim1[2], 'cng': lim2[2], 'eturna': 0, 'tgdi': 0},
                    'day4': {'nova': lim1[3], 'cng': lim2[3], 'eturna': 0, 'tgdi': 0},
                },
                'colors': color_counts,
                'vc_lists': vc_lists
            }
        else:
            wb.summary_counts = {
                'picked': {'nova': 0, 'cng': 0, 'eturna': picked_e, 'tgdi': picked_t},
                'day1': {'nova': 0, 'cng': 0, 'eturna': d1_c1, 'tgdi': d1_c2},
                'day2': {'nova': 0, 'cng': 0, 'eturna': d2_c1, 'tgdi': d2_c2},
                'day3': {'nova': 0, 'cng': 0, 'eturna': d3_c1, 'tgdi': d3_c2},
                'day4': {'nova': 0, 'cng': 0, 'eturna': d4_c1, 'tgdi': d4_c2},
                'targets': {
                    'day1': {'nova': 0, 'cng': 0, 'eturna': lim1[0], 'tgdi': lim2[0]},
                    'day2': {'nova': 0, 'cng': 0, 'eturna': lim1[1], 'tgdi': lim2[1]},
                    'day3': {'nova': 0, 'cng': 0, 'eturna': lim1[2], 'tgdi': lim2[2]},
                    'day4': {'nova': 0, 'cng': 0, 'eturna': lim1[3], 'tgdi': lim2[3]},
                },
                'colors': color_counts,
                'vc_lists': vc_lists
            }
        
        from openpyxl.styles import PatternFill
        fills = {
            'light_green': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
            'light_blue': PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),
            'light_yellow': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'light_pink': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        }
        
        # Write Day 1 Sequence
        for idx, item in enumerate(day1_list):
            vc, color_key = item
            row_idx = idx + 3
            if row_idx <= day1_ws.max_row:
                day1_ws.cell(row=row_idx, column=2).value = vc
                desc_val, col_val, svc_val = vc_lookup.get(vc, (None, None, ""))
                day1_ws.cell(row=row_idx, column=3).value = desc_val
                day1_ws.cell(row=row_idx, column=4).value = col_val
                if day1_ws.cell(row=row_idx, column=1).value is None:
                    day1_ws.cell(row=row_idx, column=1).value = idx + 1
                if color_key in fills:
                    day1_ws.cell(row=row_idx, column=2).fill = fills[color_key]
                    
        # Write Day 2, Day 3, Day 4 to biw Next 2 days sheet
        if 'biw Next 2 days' in wb.sheetnames:
            biw_ws = wb['biw Next 2 days']
            # Day 2 -> B (2)
            for idx, item in enumerate(day2_list):
                vc, color_key = item
                row_idx = idx + 2
                if row_idx <= biw_ws.max_row:
                    biw_ws.cell(row=row_idx, column=2).value = vc
                    if biw_ws.cell(row=row_idx, column=1).value is None:
                        biw_ws.cell(row=row_idx, column=1).value = idx + 1
                    if color_key in fills:
                        biw_ws.cell(row=row_idx, column=2).fill = fills[color_key]
            # Day 3 -> C (3)
            for idx, item in enumerate(day3_list):
                vc, color_key = item
                row_idx = idx + 2
                if row_idx <= biw_ws.max_row:
                    biw_ws.cell(row=row_idx, column=3).value = vc
                    if biw_ws.cell(row=row_idx, column=1).value is None:
                        biw_ws.cell(row=row_idx, column=1).value = idx + 1
                    if color_key in fills:
                        biw_ws.cell(row=row_idx, column=3).fill = fills[color_key]
            # Day 4 -> D (4)
            for idx, item in enumerate(day4_list):
                vc, color_key = item
                row_idx = idx + 2
                if row_idx <= biw_ws.max_row:
                    biw_ws.cell(row=row_idx, column=4).value = vc
                    if biw_ws.cell(row=row_idx, column=1).value is None:
                        biw_ws.cell(row=row_idx, column=1).value = idx + 1
                    if color_key in fills:
                        biw_ws.cell(row=row_idx, column=4).fill = fills[color_key]
    else:
        # Standard Phase 4 mode: only remaining Float VCs in Day 1 Sequence
        from openpyxl.styles import PatternFill
        lg_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        for idx, vc in enumerate(remaining_vcs):
            row_idx = idx + 3
            if row_idx <= day1_ws.max_row:
                day1_ws.cell(row=row_idx, column=2).value = vc
                desc_val, col_val, svc_val = vc_lookup.get(vc, (None, None, ""))
                day1_ws.cell(row=row_idx, column=3).value = desc_val
                day1_ws.cell(row=row_idx, column=4).value = col_val
                if day1_ws.cell(row=row_idx, column=1).value is None:
                    day1_ws.cell(row=row_idx, column=1).value = idx + 1
                day1_ws.cell(row=row_idx, column=2).fill = lg_fill
                    
    return wb
